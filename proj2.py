
import streamlit as st
from streamlit_option_menu import option_menu
# st.beta_set_page_config(page_title='project 2', page_icon=':smiley:')
import datetime
import pytz
from PIL import Image
import pandas as pd
from pathlib import Path
import os
import streamlit.components.v1 as stc
import pandas as pd
import openpyxl
import math
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, fills
import glob
import os
from datetime import datetime
start_time = datetime.now()

# Help


def tut_07(file, mod):

    df1 = pd.read_excel(file)  # reading the input file
    avg_u = df1['U'].mean()  # Calculating average of U,V,W
    avg_v = df1['V'].mean()
    avg_w = df1['W'].mean()

    df1["U_Avg"] = ''  # Creating average for coloumns U,V,W
    df1["V_Avg"] = ''
    df1["W_Avg"] = ''
    # assigning the values to respectivley Coloumn
    df1.iloc[0, 4] = round(avg_u, 3)
    df1.iloc[0, 5] = round(avg_v, 3)
    df1.iloc[0, 6] = round(avg_w, 3)

    # Creating new coloumns with Header U',V',W'
    df1["U'=U - U avg"] = round(df1["U"]-avg_u, 3)
    df1["V'=V - V avg"] = round(df1["V"]-avg_v, 3)
    df1["W'=W - W avg"] = round(df1["W"]-avg_w, 3)

    # df1.to_csv('octant_output.csv')

    #######          Data PreProcessing     ###########

    df1["Octant"] = ''  # Creatig a empty Column with Header as Octant

    l = len(df1)  # length of DataFrame = 29745

    # creating octant column ,and Identifying the octant value for each triple(U_1,V_1,W_1)
    for i in range(0, l):

        if (df1.loc[i, "U'=U - U avg"] >= 0 and df1.loc[i, "V'=V - V avg"] >= 0 and df1.loc[i, "W'=W - W avg"] >= 0):
            df1.loc[i, "Octant"] = "+1"  # for +1

        if (df1.loc[i, "U'=U - U avg"] >= 0 and df1.loc[i, "V'=V - V avg"] >= 0 and df1.loc[i, "W'=W - W avg"] < 0):
            df1.loc[i, "Octant"] = "-1"  # for -1

        if (df1.loc[i, "U'=U - U avg"] < 0 and df1.loc[i, "V'=V - V avg"] >= 0 and df1.loc[i, "W'=W - W avg"] >= 0):
            df1.loc[i, "Octant"] = "+2"  # for +2

        if (df1.loc[i, "U'=U - U avg"] < 0 and df1.loc[i, "V'=V - V avg"] >= 0 and df1.loc[i, "W'=W - W avg"] < 0):
            df1.loc[i, "Octant"] = "-2"  # for -2

        if (df1.loc[i, "U'=U - U avg"] < 0 and df1.loc[i, "V'=V - V avg"] < 0 and df1.loc[i, "W'=W - W avg"] >= 0):
            df1.loc[i, "Octant"] = "+3"  # for +3

        if (df1.loc[i, "U'=U - U avg"] < 0 and df1.loc[i, "V'=V - V avg"] < 0 and df1.loc[i, "W'=W - W avg"] < 0):
            df1.loc[i, "Octant"] = "-3"  # for -3

        if (df1.loc[i, "U'=U - U avg"] >= 0 and df1.loc[i, "V'=V - V avg"] < 0 and df1.loc[i, "W'=W - W avg"] >= 0):
            df1.loc[i, "Octant"] = "+4"  # for +4

        if (df1.loc[i, "U'=U - U avg"] >= 0 and df1.loc[i, "V'=V - V avg"] < 0 and df1.loc[i, "W'=W - W avg"] < 0):
            df1.loc[i, "Octant"] = "-4"  # for -4

            ######  Octant Identification  ########

    # creating empty Column without header and assigned "User input" to row 3
    df1[""] = "  "
    df1[" "] = " "
    df1.iloc[0, 12] = "Mod "+str(mod)

    # creating a Coloumn with header as Octant ID
    df1["Octant ID"] = " "
    df1.loc[0, "Octant ID"] = "Overall Octant"

    # oct_count stores a count of unique elements i.e. count of +1,-1,+2,-2,+3,-4,+4
    oct_count = df1['Octant'].value_counts()

    arr = ["+1", "-1", "+2", "-2", "+3", "-3",
           "+4", "-4"]  # cretaed for reference

    oct_cnt = {}  # for storing octant count as key and and coreesponding octant value as value in dict
    for i in range(8):
        s = arr[i]
        # appending the overall count of octant and octant value in dict i.e for Ex:("+1",2610)
        oct_cnt.update({s: oct_count[s]})
        # And assigning a count values to respectively Coloumns
        df1.loc[0, s] = oct_count[s]

    # print(oct_cnt) #{2610: '+1', 4603: '-1', 4855: '+2', 2798: '-2', 4548: '+3', 2784: '-3', 2769: '+4', 4778: '-4'}
    # sorting the dict by keys
    # print(sortedbykey) {2610: '+1', 2769: '+4', 2784: '-3', 2798: '-2', 4548: '+3', 4603: '-1', 4778: '-4', 4855: '+2'}
    # storing the sorted values in a list
    # print(sortedbyval_lst)['+1', '+4', '-3', '-2', '+3', '-1', '-4', '+2']

    # sorting the dict by values
    sortedbyval = {k: v for k, v in sorted(
        oct_cnt.items(), key=lambda item: item[1])}
    # storing the sorted keys in a list
    sortedbyval_lst = list(sortedbyval.keys())

    octant_name_id_mapping = {"1": "Internal outward interaction", "-1": "External outward interaction", "2": "External Ejection",
                              "-2": "Internal Ejection", "3": "External inward interaction", "-3": "Internal inward interaction", "4": "Internal sweep", "-4": "External sweep"}

    df1["Rank 1"] = ''  # created empty columns
    df1["Rank 2"] = ''
    df1["Rank 3"] = ''
    df1["Rank 4"] = ''
    df1["Rank 5"] = ''
    df1["Rank 6"] = ''
    df1["Rank 7"] = ''
    df1["Rank 8"] = ''
    df1["Rank1 Octant ID"] = " "

    dic_rank = {"+1": "Rank 1", "-1": "Rank 2", "+2": "Rank 3", "-2": "Rank 4",
                "+3": "Rank 5", "-3": "Rank 6", "+4": "Rank 7", "-4": "Rank 8"}  # for reference

    # i=0
    for i in range(8):
        df1.loc[0, dic_rank[sortedbyval_lst[i]]] = 8 - \
            i  # appending the octant ranks of octants
        if (8-i == 1):
            # appending the highest rank octant and its corresponding octant name
            df1.loc[0, "Rank1 Octant ID"] = sortedbyval_lst[i]
            df1.loc[0, "Rank1 Octant Name"] = octant_name_id_mapping[str(
                int(df1.loc[0, "Rank1 Octant ID"]))]

            ###########   Added Some Columns And Rows for MOD Count   ##########

    x = 0  # for findind octant values for MOD ranges
    t = 1  # for row pointer

    count_rank_mod = [0]*8  # Count of rank mod values
    while (x < l):

        d1 = {"+1": 0, "-1": 1, "+2": 2, "-2": 3, "+3": 4,
              "-3": 5, "+4": 6, "-4": 7}  # creating a dictionary for reference

        # count values of each octant is stored for MOD ranges
        oct_cnt_mod = [0]*8

        oct_cnt = {}  # for storing octant count as key and and coreesponding octant value as value in dict
        for i in range(x, x+mod, 1):

            if (i >= l):
                break  # bound check
            s3 = df1.at[i, "Octant"]
            # incrementing by one of count values of corresponding octants
            oct_cnt_mod[d1[s3]] += 1

        i = 0
        for i in range(8):
            s = arr[i]
            # assigning overall count of octants in each interval
            df1.loc[t, s] = oct_cnt_mod[i]
            # appending the overall count of octant and octant value in dict
            oct_cnt.update({s: oct_cnt_mod[i]})

        # sorting the dict by values
        sortedbyval = {k: v for k, v in sorted(
            oct_cnt.items(), key=lambda item: item[1])}
        # storing the sorted keys in a list
        sortedbyval_lst = list(sortedbyval.keys())

        i = 0
        for i in range(8):
            df1.loc[t, dic_rank[sortedbyval_lst[i]]] = 8 - \
                i  # appending the octant ranks of octants
            if (8-i == 1):
                # appending the highest rank octant and its corresponding octant name
                df1.loc[t, "Rank1 Octant ID"] = sortedbyval_lst[i]
                df1.loc[t, "Rank1 Octant Name"] = octant_name_id_mapping[str(
                    int(df1.loc[t, "Rank1 Octant ID"]))]
                # incrementing by one of corresponding octant
                count_rank_mod[d1[sortedbyval_lst[i]]] += 1

        if ((x+mod) > l):  # Writing MOD ranges in Octant ID Coloumn
            df1.loc[t, "Octant ID"] = str(
                x)+"-"+str(l-1)  # for last index(i.e) 2744
        else:
            df1.loc[t, "Octant ID"] = str(x)+"-"+str(x+mod-1)

        x += mod
        t += 1

        ################ Octant Count Based on Mod Values  ######################

    t += 3
    df1.loc[t, "+1"] = "Octant ID"
    df1.loc[t, "-1"] = "Octant Name"
    df1.loc[t, "+2"] = "Count of Rank1 of Mod Values"
    t += 1
    i = 0
    for ID, name in octant_name_id_mapping.items():  # iterating through a dict
        # appending the Octant IDs, Octant Name ,and count of Rank1 of mod values
        df1.loc[t, "+1"] = int(ID)
        df1.loc[t, "-1"] = name
        df1.loc[t, "+2"] = count_rank_mod[i]
        t += 1
        i += 1

        ############### tut 5 ###############
        ################ Octant Count Based on Mod Values  ######################
    df1["  "] = ""
    df1["   "] = " "
    df1.iloc[0, 33] = "From"
    df1.loc["Octant #"] = " "
    arr = [" +1", " -1", " +2", " -2", " +3", " -3", " +4", " -4"]

    j = 0
    for i in range(0, 8):
        df1.loc[i, "Octant #"] = arr[j]  # updating Octant ID column
        j += 1

    j = 0
    for j in range(0, 8):
        s1 = arr[j]  # verifing the count of octants
        df1[s1] = " "

    t1 = 0
    t2 = 1
    d1 = {"+1": 0, "-1": 1, "+2": 2, "-2": 3,
          "+3": 4, "-3": 5, "+4": 6, "-4": 7}
    d2 = {"+1": " +1", "-1": " -1", "+2": " +2", "-2": " -2",
          "+3": " +3", "-3": " -3", "+4": " +4", "-4": " -4"}

    while (1):
        if (t2 == l):
            break
        s1 = df1.at[t1, "Octant"]  # From
        s2 = df1.at[t2, "Octant"]  # To
        # print(df1.loc[d1[s1], d2[s2]])
        if (df1.loc[d1[s1], d2[s2]] == " "):  # checking if cell is empty/null
            df1.loc[d1[s1], d2[s2]] = 1  # adding one
        else:
            # increamenting the count by one and updating it to coloumn
            df1.loc[d1[s1], d2[s2]] = int(df1.loc[d1[s1], d2[s2]]) + 1
        t1 += 1
        t2 += 1

    t = 7
    x = 0
    while (x < l):
        t += 4
        df1.loc[t, "Octant #"] = "Mod Transition Count"
        if ((x+mod) > l):  # Writing MOD ranges in Octant ID Coloumn
            # for last index(i.e) 2744
            df1.loc[t+1, "Octant #"] = str(x)+"-"+str(l-1)
        else:
            df1.loc[t+1, "Octant #"] = str(x)+"-"+str(x+mod-1)
        df1.loc[t+1, " +1"] = "To"
        t += 2
        arr = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
        df1.loc[t, "Octant #"] = "Octant #"
        df1.iloc[t+1, 33] = "From"
        h = df1.columns  # h stores column labels
        # header name in index format(integer) (here ,y=13)
        y = h.get_loc(" +1")
        j = 0
        for i in range(y, y+8):  # updating a row
            df1.iloc[t, i] = arr[j]
            j += 1

        j = 0
        for i in range(t+1, t+9):  # updating Coloumn
            df1.loc[i, "Octant #"] = arr[j]
            j += 1

        for i in range(x, x+mod):  # each interval

            if (i == l-1):
                break
            s1 = df1.at[i, "Octant"]  # From
            s2 = df1.at[i+1, "Octant"]  # To

            if (df1.loc[t+d1[s1]+1, d2[s2]] == " "):  # checking if cell is empty/null
                df1.loc[t+d1[s1]+1, d2[s2]] = 1  # adding one
            else:
                # increamenting the count by one and updating it to coloumn
                df1.loc[t+d1[s1]+1, d2[s2]
                        ] = int(df1.loc[t+d1[s1]+1, d2[s2]]) + 1
        t += 8
        x += mod

        ############### tut 2 ###############
        ##################
    df1["    "] = " "
    # Creating empty column with Octant as a header
    df1["Octant ##"] = " "
    arr = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
    for i in range(8):
        # appending values in octant column
        df1.loc[i, "Octant ##"] = arr[i]

    df1["Longest Subsequence Length"] = " "
    df1["Count"] = " "

    x = 0
    # Longest subsequence length for respectively octant values #initlizing a max_count with all zeroes  #initlizing a max_count with all zeroes
    max_count = [0]*8

    # for count of LSL for respectively octant values # initlizing a max_count with all zeroes #initlizing a LSL_count with all zeroes
    LSL_count = [0]*8
    d1 = {"+1": 0, "-1": 1, "+2": 2, "-2": 3, "+3": 4,
          "-3": 5, "+4": 6, "-4": 7}  # creating a dictionary

    # Creating an empty 2d list of size of 8
    # where each list stores the upper range(Time Range) value of thier respectively Octants
    time_range = []
    for i in range(8):
        time_range.append([])

    while (x < l):
        s1 = df1.at[x, "Octant"]
        count = 0
        j = x
        while (1):  # counting length of sequence
            # breaking if next element is not equal to s1
            if (j >= l or df1.at[j, "Octant"] != s1):
                break
            count += 1
            j += 1

        x += count
        temp = max_count[d1[s1]]

        # updating a maximum count of value if current count is greater the current max
        max_count[d1[s1]] = max(max_count[d1[s1]], count)

        if (count > temp):
            # Reassigning the values of LSL count to one
            LSL_count[d1[s1]] = 1
            # if list is empty appending  Upper range Value
            if (len(time_range[d1[s1]]) == 0):
                time_range[d1[s1]].append(j-1)

            else:
                time_range[d1[s1]].clear()  # Clearing the list
                # appending a curent upper range value to the same clered octant list
                time_range[d1[s1]].append(j-1)

        if (count == temp):
            # incremneting the count of LSL by one
            LSL_count[d1[s1]] += 1
            # appending to the pre-existing(non-empty) list having same LSL of respective Octant
            time_range[d1[s1]].append(j-1)

            ############### tut 2 ###############

    for i in range(8):
        # updating Longest subsequence length for respectively octant values
        df1.loc[i, "Longest Subsequence Length"] = max_count[i]

    max_l_cnt = 0
    for j in range(8):
        # updating count of LSL for respectively octant values
        df1.loc[j, "Count"] = LSL_count[j]
        max_l_cnt += LSL_count[j]

    df1["     "] = " "  # Empty Column without Header
    df1["Octant ####"] = " "  # Empty Column
    df1[" Longest Subsequence Length"] = " "  # Empty Column
    df1[" Count"] = " "  # Empty Column
    # print(time_range) # time_range = [[10945], [14645, 18174, 19131], [16990], [29321], [16217], [677], [29219], [28059]]

    t = 0  # row pointer
    for i in range(8):
        df1.loc[t, "Octant ####"] = arr[i]  # Updating Octant Values
        # Updating LSL of Octants
        df1.loc[t, " Longest Subsequence Length"] = max_count[i]
        # updating count of LSl of Octants
        df1.loc[t, " Count"] = LSL_count[i]
        t += 1  # t points to next row
        df1.loc[t, "Octant ####"] = "Time"
        df1.loc[t, " Longest Subsequence Length"] = "From"
        df1.loc[t, " Count"] = "To"

        t += 1  # t points to next row
        for j in range(LSL_count[i]):
            # Appending lower range # From
            df1.loc[t, " Longest Subsequence Length"] = 0.01 * \
                ((time_range[d1[arr[i]]][j])-(max_count[i]-1))
            # Appending Upper range #To
            df1.loc[t, " Count"] = 0.01*time_range[d1[arr[i]]][j]
            t += 1

            ############### tut 3 and 4 ###############

    # taking 1st name of input file for naming of output
    # inp = lst_files[sat].replace(
    # 	'.xlsx', " cm_vel_octant_analysis_mod_"+str(mod)+".xlsx")
    # # changing direc to ouput file to save output files
    # os.chdir(path_output)
    df1.to_excel("output.xlsx", index=False)  # updating dataframe into excel

    ########################################################################################
    # oprating on openpyxl for borders and colouring
    wb = openpyxl.load_workbook("output.xlsx")
    ws = wb['Sheet1']

    # for borders a cell
    thin_border = Border(left=Side(border_style='thin', color='FF000000'),
                         right=Side(border_style='thin',
                                    color='FF000000'),
                         top=Side(border_style='thin',
                                  color='FF000000'),
                         bottom=Side(border_style='thin',
                                     color='FF000000')
                         )

    thick_border = Border(left=Side(border_style='thin', color='FF000000'),
                          right=Side(border_style='thin',
                                     color='FF000000'),
                          top=Side(border_style='thin',
                                   color='FF000000'),
                          bottom=Side(border_style='medium',
                                      color='FF000000')
                          )

    # for colouring a cell
    fill_cell = PatternFill(
        fill_type=fills.FILL_SOLID, start_color='00FFFF00', end_color='00FFFF00')

    # For Overall octant count and Rank of Octant
    # define size of the table     
    row_num = math.ceil(l/mod)+2
    col_num = 19
    # location of the Table
    row_loc = 1
    col_loc = 14

    for i in range(row_loc, row_loc+row_num):
        for j in range(col_loc, col_num+col_loc):
            ws.cell(row=i, column=j).border = thin_border
            if ((ws.cell(row=i, column=j).value == 1)):  # adding colour for rank ! values
                ws.cell(row=i, column=j).fill = fill_cell
            if i == row_loc+row_num-1:
                ws.cell(row=i, column=j).border = thick_border

    # define size of the table  # for count of Rank1 values
    row_num = 9
    col_num = 3
    # location of the Table
    row_loc = math.ceil(l/mod)+6  # (2+1+4)
    col_loc = 15

    for i in range(row_loc, row_loc+row_num):
        for j in range(col_loc, col_num+col_loc):
            ws.cell(row=i, column=j).border = thin_border
            if i == row_loc+row_num-1:
                ws.cell(row=i, column=j).border = thick_border

    # define size of the table  # for overall trasition and mod transition
    row_num = 9
    col_num = 9
    # location of the Table
    row_loc = 1
    col_loc = 35

    # Number of Tables
    Table_num = math.ceil(l/mod)+1
    dis = 5  # distance between the tables

    for _ in range(Table_num):
        k = 0
        for i in range(row_loc, row_loc+row_num):

            if (i > row_loc):
                ws.cell(row=i, column=col_loc+k).fill = fill_cell
            for j in range(col_loc, col_num+col_loc):
                if ((ws.cell(row=i, column=j).value == " ")):
                    ws.cell(row=i, column=j).value=0
                ws.cell(row=i, column=j).border = thin_border
                if i == row_loc+row_num-1:
                    ws.cell(row=i, column=j).border = thick_border
            k += 1

        row_loc = row_loc+row_num+dis

    # define size of the table   ##For Time Ranges of Octant's LSL
    row_num = 9
    col_num = 3
    # location of the Table
    row_loc = 1
    col_loc = 45

    for i in range(row_loc, row_loc+row_num):
        for j in range(col_loc, col_num+col_loc):
            ws.cell(row=i, column=j).border = thin_border
            if i == row_loc+row_num-1:
                ws.cell(row=i, column=j).border = thick_border

    row_num = l  # For Time Ranges of Octant's LSL
    col_num = 3
    # location of the Table
    row_loc = 1
    col_loc = 49

    for i in range(row_loc, row_loc+row_num):
        # print(ws.cell(row=i, column=50).value)
        if (ws.cell(row=i, column=50).value == " "):  # breakig if values is None( empty cell)
            break
        for j in range(col_loc, col_num+col_loc):
            ws.cell(row=i, column=j).border = thin_border
            if i == row_loc+row_num-1:
                ws.cell(row=i, column=j).border = thick_border

    curr_time = datetime.now(pytz.timezone("Asia/Kolkata")).strftime('%Y-%m-%d %H:%M:%S:%f')
    out_nam = (file.name).replace('.xlsx', "_mod_"+str(mod)+"_"+curr_time + ".xlsx")
    out_nam = out_nam.replace(" ", "-").replace(":", "-")

    # st.download_button(label='📥 Download Current Result',
    #                             data=wb ,
    #                             file_name= out_nam)
    wb.save(out_nam)  # saving the file


def tut_007(path_input, mod,path_output):

    try:
        # path = 'G:\CS384\2001EE19_2022\tut07\output' input from function
        # Check whether the specified
        # path exists or not
        
        isExist = os.path.isdir(path_output)
        if (not (isExist)):
            os.mkdir(path_output)

        os.chdir(path_input)
        lst_files = glob.glob('*.xlsx')
        sat = 0
        for file in lst_files:  # iterating thorugh files in input folder

            # again we changing the dir to input from output
            os.chdir(path_input)
            df1 = pd.read_excel(file)  # reading the input file
            avg_u = df1['U'].mean()  # Calculating average of U,V,W
            avg_v = df1['V'].mean()
            avg_w = df1['W'].mean()

            df1["U_Avg"] = ''  # Creating average for coloumns U,V,W
            df1["V_Avg"] = ''
            df1["W_Avg"] = ''
            # assigning the values to respectivley Coloumn
            df1.iloc[0, 4] = round(avg_u, 3)
            df1.iloc[0, 5] = round(avg_v, 3)
            df1.iloc[0, 6] = round(avg_w, 3)

            # Creating new coloumns with Header U',V',W'
            df1["U'=U - U avg"] = round(df1["U"]-avg_u, 3)
            df1["V'=V - V avg"] = round(df1["V"]-avg_v, 3)
            df1["W'=W - W avg"] = round(df1["W"]-avg_w, 3)

            # df1.to_csv('octant_output.csv')

            #######          Data PreProcessing     ###########

            df1["Octant"] = ''  # Creatig a empty Column with Header as Octant

            l = len(df1)  # length of DataFrame = 29745

            # creating octant column ,and Identifying the octant value for each triple(U_1,V_1,W_1)
            for i in range(0, l):

                if (df1.loc[i, "U'=U - U avg"] >= 0 and df1.loc[i, "V'=V - V avg"] >= 0 and df1.loc[i, "W'=W - W avg"] >= 0):
                    df1.loc[i, "Octant"] = "+1"  # for +1

                if (df1.loc[i, "U'=U - U avg"] >= 0 and df1.loc[i, "V'=V - V avg"] >= 0 and df1.loc[i, "W'=W - W avg"] < 0):
                    df1.loc[i, "Octant"] = "-1"  # for -1

                if (df1.loc[i, "U'=U - U avg"] < 0 and df1.loc[i, "V'=V - V avg"] >= 0 and df1.loc[i, "W'=W - W avg"] >= 0):
                    df1.loc[i, "Octant"] = "+2"  # for +2

                if (df1.loc[i, "U'=U - U avg"] < 0 and df1.loc[i, "V'=V - V avg"] >= 0 and df1.loc[i, "W'=W - W avg"] < 0):
                    df1.loc[i, "Octant"] = "-2"  # for -2

                if (df1.loc[i, "U'=U - U avg"] < 0 and df1.loc[i, "V'=V - V avg"] < 0 and df1.loc[i, "W'=W - W avg"] >= 0):
                    df1.loc[i, "Octant"] = "+3"  # for +3

                if (df1.loc[i, "U'=U - U avg"] < 0 and df1.loc[i, "V'=V - V avg"] < 0 and df1.loc[i, "W'=W - W avg"] < 0):
                    df1.loc[i, "Octant"] = "-3"  # for -3

                if (df1.loc[i, "U'=U - U avg"] >= 0 and df1.loc[i, "V'=V - V avg"] < 0 and df1.loc[i, "W'=W - W avg"] >= 0):
                    df1.loc[i, "Octant"] = "+4"  # for +4

                if (df1.loc[i, "U'=U - U avg"] >= 0 and df1.loc[i, "V'=V - V avg"] < 0 and df1.loc[i, "W'=W - W avg"] < 0):
                    df1.loc[i, "Octant"] = "-4"  # for -4

                    ######  Octant Identification  ########

            # creating empty Column without header and assigned "User input" to row 3
            df1[""] = "  "
            df1[" "] = " "
            df1.iloc[0, 12] = "Mod "+str(mod)

            # creating a Coloumn with header as Octant ID
            df1["Octant ID"] = " "
            df1.loc[0, "Octant ID"] = "Overall Octant"

            # oct_count stores a count of unique elements i.e. count of +1,-1,+2,-2,+3,-4,+4
            oct_count = df1['Octant'].value_counts()

            arr = ["+1", "-1", "+2", "-2", "+3", "-3",
                   "+4", "-4"]  # cretaed for reference

            oct_cnt = {}  # for storing octant count as key and and coreesponding octant value as value in dict
            for i in range(8):
                s = arr[i]
                # appending the overall count of octant and octant value in dict i.e for Ex:("+1",2610)
                oct_cnt.update({s: oct_count[s]})
                # And assigning a count values to respectively Coloumns
                df1.loc[0, s] = oct_count[s]

            # print(oct_cnt) #{2610: '+1', 4603: '-1', 4855: '+2', 2798: '-2', 4548: '+3', 2784: '-3', 2769: '+4', 4778: '-4'}
            # sorting the dict by keys
            # print(sortedbykey) {2610: '+1', 2769: '+4', 2784: '-3', 2798: '-2', 4548: '+3', 4603: '-1', 4778: '-4', 4855: '+2'}
            # storing the sorted values in a list
            # print(sortedbyval_lst)['+1', '+4', '-3', '-2', '+3', '-1', '-4', '+2']

            # sorting the dict by values
            sortedbyval = {k: v for k, v in sorted(
                oct_cnt.items(), key=lambda item: item[1])}
            # storing the sorted keys in a list
            sortedbyval_lst = list(sortedbyval.keys())

            octant_name_id_mapping = {"1": "Internal outward interaction", "-1": "External outward interaction", "2": "External Ejection",
                                      "-2": "Internal Ejection", "3": "External inward interaction", "-3": "Internal inward interaction", "4": "Internal sweep", "-4": "External sweep"}

            df1["Rank 1"] = ''  # created empty columns
            df1["Rank 2"] = ''
            df1["Rank 3"] = ''
            df1["Rank 4"] = ''
            df1["Rank 5"] = ''
            df1["Rank 6"] = ''
            df1["Rank 7"] = ''
            df1["Rank 8"] = ''
            df1["Rank1 Octant ID"] = " "

            dic_rank = {"+1": "Rank 1", "-1": "Rank 2", "+2": "Rank 3", "-2": "Rank 4",
                        "+3": "Rank 5", "-3": "Rank 6", "+4": "Rank 7", "-4": "Rank 8"}  # for reference

            # i=0
            for i in range(8):
                df1.loc[0, dic_rank[sortedbyval_lst[i]]] = 8 - \
                    i  # appending the octant ranks of octants
                if (8-i == 1):
                    # appending the highest rank octant and its corresponding octant name
                    df1.loc[0, "Rank1 Octant ID"] = sortedbyval_lst[i]
                    df1.loc[0, "Rank1 Octant Name"] = octant_name_id_mapping[str(
                        int(df1.loc[0, "Rank1 Octant ID"]))]

                    ###########   Added Some Columns And Rows for MOD Count   ##########

            x = 0  # for findind octant values for MOD ranges
            t = 1  # for row pointer

            count_rank_mod = [0]*8  # Count of rank mod values
            while (x < l):

                d1 = {"+1": 0, "-1": 1, "+2": 2, "-2": 3, "+3": 4,
                      "-3": 5, "+4": 6, "-4": 7}  # creating a dictionary for reference

                # count values of each octant is stored for MOD ranges
                oct_cnt_mod = [0]*8

                oct_cnt = {}  # for storing octant count as key and and coreesponding octant value as value in dict
                for i in range(x, x+mod, 1):

                    if (i >= l):
                        break  # bound check
                    s3 = df1.at[i, "Octant"]
                    # incrementing by one of count values of corresponding octants
                    oct_cnt_mod[d1[s3]] += 1

                i = 0
                for i in range(8):
                    s = arr[i]
                    # assigning overall count of octants in each interval
                    df1.loc[t, s] = oct_cnt_mod[i]
                    # appending the overall count of octant and octant value in dict
                    oct_cnt.update({s: oct_cnt_mod[i]})

                # sorting the dict by values
                sortedbyval = {k: v for k, v in sorted(
                    oct_cnt.items(), key=lambda item: item[1])}
                # storing the sorted keys in a list
                sortedbyval_lst = list(sortedbyval.keys())

                i = 0
                for i in range(8):
                    df1.loc[t, dic_rank[sortedbyval_lst[i]]] = 8 - \
                        i  # appending the octant ranks of octants
                    if (8-i == 1):
                        # appending the highest rank octant and its corresponding octant name
                        df1.loc[t, "Rank1 Octant ID"] = sortedbyval_lst[i]
                        df1.loc[t, "Rank1 Octant Name"] = octant_name_id_mapping[str(int(df1.loc[t, "Rank1 Octant ID"]))]
                        # incrementing by one of corresponding octant
                        count_rank_mod[d1[sortedbyval_lst[i]]] += 1

                if ((x+mod) > l):  # Writing MOD ranges in Octant ID Coloumn
                    df1.loc[t, "Octant ID"] = str(
                        x)+"-"+str(l-1)  # for last index(i.e) 2744
                else:
                    df1.loc[t, "Octant ID"] = str(x)+"-"+str(x+mod-1)

                x += mod
                t += 1

                ################ Octant Count Based on Mod Values  ######################

            t += 3
            df1.loc[t, "+1"] = "Octant ID"
            df1.loc[t, "-1"] = "Octant Name"
            df1.loc[t, "+2"] = "Count of Rank1 of Mod Values"
            t += 1
            i = 0
            for ID, name in octant_name_id_mapping.items():  # iterating through a dict
                # appending the Octant IDs, Octant Name ,and count of Rank1 of mod values
                df1.loc[t, "+1"] = int(ID)
                df1.loc[t, "-1"] = name
                df1.loc[t, "+2"] = count_rank_mod[i]
                t += 1
                i += 1

                ############### tut 5 ###############
                ################ Octant Count Based on Mod Values  ######################
            df1["  "] = ""
            df1["   "] = " "
            df1.iloc[0, 33] = "From"
            df1.loc["Octant #"] = " "
            arr = [" +1", " -1", " +2", " -2", " +3", " -3", " +4", " -4"]

            j = 0
            for i in range(0, 8):
                df1.loc[i, "Octant #"] = arr[j]  # updating Octant ID column
                j += 1

            j = 0
            for j in range(0, 8):
                s1 = arr[j]  # verifing the count of octants
                df1[s1] = " "

            t1 = 0
            t2 = 1
            d1 = {"+1": 0, "-1": 1, "+2": 2, "-2": 3,
                  "+3": 4, "-3": 5, "+4": 6, "-4": 7}
            d2 = {"+1": " +1", "-1": " -1", "+2": " +2", "-2": " -2",
                  "+3": " +3", "-3": " -3", "+4": " +4", "-4": " -4"}

            while (1):
                if (t2 == l):
                    break
                s1 = df1.at[t1, "Octant"]  # From
                s2 = df1.at[t2, "Octant"]  # To
                # print(df1.loc[d1[s1], d2[s2]])
                if (df1.loc[d1[s1], d2[s2]] == " "):  # checking if cell is empty/null
                    df1.loc[d1[s1], d2[s2]] = 1  # adding one
                else:
                    # increamenting the count by one and updating it to coloumn
                    df1.loc[d1[s1], d2[s2]] = int(df1.loc[d1[s1], d2[s2]]) + 1
                t1 += 1
                t2 += 1

            t = 7
            x = 0
            while (x < l):
                t += 4
                df1.loc[t, "Octant #"] = "Mod Transition Count"
                if ((x+mod) > l):  # Writing MOD ranges in Octant ID Coloumn
                    # for last index(i.e) 2744
                    df1.loc[t+1, "Octant #"] = str(x)+"-"+str(l-1)
                else:
                    df1.loc[t+1, "Octant #"] = str(x)+"-"+str(x+mod-1)
                df1.loc[t+1, " +1"] = "To"
                t += 2
                arr = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
                df1.loc[t, "Octant #"] = "Octant #"
                df1.iloc[t+1, 33] = "From"
                h = df1.columns  # h stores column labels
                # header name in index format(integer) (here ,y=13)
                y = h.get_loc(" +1")
                j = 0
                for i in range(y, y+8):  # updating a row
                    df1.iloc[t, i] = arr[j]
                    j += 1

                j = 0
                for i in range(t+1, t+9):  # updating Coloumn
                    df1.loc[i, "Octant #"] = arr[j]
                    j += 1

                for i in range(x, x+mod):  # each interval

                    if (i == l-1):
                        break
                    s1 = df1.at[i, "Octant"]  # From
                    s2 = df1.at[i+1, "Octant"]  # To

                    if (df1.loc[t+d1[s1]+1, d2[s2]] == " "):  # checking if cell is empty/null
                        df1.loc[t+d1[s1]+1, d2[s2]] = 1  # adding one
                    else:
                        # increamenting the count by one and updating it to coloumn
                        df1.loc[t+d1[s1]+1, d2[s2]
                                ] = int(df1.loc[t+d1[s1]+1, d2[s2]]) + 1
                t += 8
                x += mod

                ############### tut 2 ###############
                ##################
            df1["    "] = " "
            # Creating empty column with Octant as a header
            df1["Octant ##"] = " "
            arr = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
            for i in range(8):
                # appending values in octant column
                df1.loc[i, "Octant ##"] = arr[i]

            df1["Longest Subsequence Length"] = " "
            df1["Count"] = " "

            x = 0
            # Longest subsequence length for respectively octant values #initlizing a max_count with all zeroes  #initlizing a max_count with all zeroes
            max_count = [0]*8

            # for count of LSL for respectively octant values # initlizing a max_count with all zeroes #initlizing a LSL_count with all zeroes
            LSL_count = [0]*8
            d1 = {"+1": 0, "-1": 1, "+2": 2, "-2": 3, "+3": 4,
                  "-3": 5, "+4": 6, "-4": 7}  # creating a dictionary

            # Creating an empty 2d list of size of 8
            # where each list stores the upper range(Time Range) value of thier respectively Octants
            time_range = []
            for i in range(8):
                time_range.append([])

            while (x < l):
                s1 = df1.at[x, "Octant"]
                count = 0
                j = x
                while (1):  # counting length of sequence
                    # breaking if next element is not equal to s1
                    if (j >= l or df1.at[j, "Octant"] != s1):
                        break
                    count += 1
                    j += 1

                x += count
                temp = max_count[d1[s1]]

                # updating a maximum count of value if current count is greater the current max
                max_count[d1[s1]] = max(max_count[d1[s1]], count)

                if (count > temp):
                    # Reassigning the values of LSL count to one
                    LSL_count[d1[s1]] = 1
                    # if list is empty appending  Upper range Value
                    if (len(time_range[d1[s1]]) == 0):
                        time_range[d1[s1]].append(j-1)

                    else:
                        time_range[d1[s1]].clear()  # Clearing the list
                        # appending a curent upper range value to the same clered octant list
                        time_range[d1[s1]].append(j-1)

                if (count == temp):
                    # incremneting the count of LSL by one
                    LSL_count[d1[s1]] += 1
                    # appending to the pre-existing(non-empty) list having same LSL of respective Octant
                    time_range[d1[s1]].append(j-1)

                    ############### tut 2 ###############

            for i in range(8):
                # updating Longest subsequence length for respectively octant values
                df1.loc[i, "Longest Subsequence Length"] = max_count[i]

            max_l_cnt = 0
            for j in range(8):
                # updating count of LSL for respectively octant values
                df1.loc[j, "Count"] = LSL_count[j]
                max_l_cnt += LSL_count[j]

            df1["     "] = " "  # Empty Column without Header
            df1["Octant ####"] = " "  # Empty Column
            df1[" Longest Subsequence Length"] = " "  # Empty Column
            df1[" Count"] = " "  # Empty Column
            # print(time_range) # time_range = [[10945], [14645, 18174, 19131], [16990], [29321], [16217], [677], [29219], [28059]]

            t = 0  # row pointer
            for i in range(8):
                df1.loc[t, "Octant ####"] = arr[i]  # Updating Octant Values
                # Updating LSL of Octants
                df1.loc[t, " Longest Subsequence Length"] = max_count[i]
                # updating count of LSl of Octants
                df1.loc[t, " Count"] = LSL_count[i]
                t += 1  # t points to next row
                df1.loc[t, "Octant ####"] = "Time"
                df1.loc[t, " Longest Subsequence Length"] = "From"
                df1.loc[t, " Count"] = "To"

                t += 1  # t points to next row
                for j in range(LSL_count[i]):
                    # Appending lower range # From
                    df1.loc[t, " Longest Subsequence Length"] = 0.01 * \
                        ((time_range[d1[arr[i]]][j])-(max_count[i]-1))
                    # Appending Upper range #To
                    df1.loc[t, " Count"] = 0.01*time_range[d1[arr[i]]][j]
                    t += 1

                    ############### tut 3 and 4 ###############

            # taking 1st name of input file for naming of output

            curr_time = datetime.now(pytz.timezone("Asia/Kolkata")).strftime('%Y-%m-%d %H:%M:%S:%f')
            out_nam = lst_files[sat].replace('.xlsx', "_mod_"+str(mod)+"_"+curr_time + ".xlsx")
            out_nam = out_nam.replace(" ", "-").replace(":", "-")    
            # changing direc to ouput file to save output files
            os.chdir(path_output)
            print(out_nam)
            df1.to_excel(out_nam, index=False)  # updating dataframe into excel

            ########################################################################################
            # oprating on openpyxl for borders and colouring
            wb = openpyxl.load_workbook(out_nam)
            ws = wb['Sheet1']

            # for borders a cell
            thin_border = Border(left=Side(border_style='thin', color='FF000000'),
                                 right=Side(border_style='thin',
                                            color='FF000000'),
                                 top=Side(border_style='thin',
                                          color='FF000000'),
                                 bottom=Side(border_style='thin',
                                             color='FF000000')
                                 )

            thick_border = Border(left=Side(border_style='thin', color='FF000000'),
                                  right=Side(border_style='thin',
                                             color='FF000000'),
                                  top=Side(border_style='thin',
                                           color='FF000000'),
                                  bottom=Side(border_style='medium',
                                              color='FF000000')
                                  )

            # for colouring a cell
            fill_cell = PatternFill(
                fill_type=fills.FILL_SOLID, start_color='00FFFF00', end_color='00FFFF00')

            # define size of the table     # For Overall octant count and Rank of Octant
            row_num = math.ceil(l/mod)+2
            col_num = 19
            # location of the Table
            row_loc = 1
            col_loc = 14

            for i in range(row_loc, row_loc+row_num):
                for j in range(col_loc, col_num+col_loc):
                    ws.cell(row=i, column=j).border = thin_border
                    if ((ws.cell(row=i, column=j).value == 1)):  # adding colour for rank ! values
                        ws.cell(row=i, column=j).fill = fill_cell
                    if i == row_loc+row_num-1:
                        ws.cell(row=i, column=j).border = thick_border

            # define size of the table  # for count of Rank1 values
            row_num = 9
            col_num = 3
            # location of the Table
            row_loc = math.ceil(l/mod)+6
            col_loc = 25

            for i in range(row_loc, row_loc+row_num):
                for j in range(col_loc, col_num+col_loc):
                    ws.cell(row=i, column=j).border = thin_border
                    if i == row_loc+row_num-1:
                        ws.cell(row=i, column=j).border = thick_border

            # define size of the table  # for overall trasition and mod transition
            row_num = 9
            col_num = 9
            # location of the Table
            row_loc = 1
            col_loc = 35

            # Number of Tables
            Table_num = math.ceil(l/mod)+1
            dis = 5  # distance between the tables

            for _ in range(Table_num):
                k = 0
                for i in range(row_loc, row_loc+row_num):

                    if (i > row_loc):
                        ws.cell(row=i, column=col_loc+k).fill = fill_cell
                    for j in range(col_loc, col_num+col_loc):
                        if ((ws.cell(row=i, column=j).value == " ")):
                            ws.cell(row=i, column=j).value=0
                        ws.cell(row=i, column=j).border = thin_border
                        if i == row_loc+row_num-1:
                            ws.cell(row=i, column=j).border = thick_border
                    k += 1

                row_loc = row_loc+row_num+dis

            # define size of the table   ##For Time Ranges of Octant's LSL
            row_num = 9
            col_num = 3
            # location of the Table
            row_loc = 1
            col_loc = 45

            for i in range(row_loc, row_loc+row_num):
                for j in range(col_loc, col_num+col_loc):
                    ws.cell(row=i, column=j).border = thin_border
                    if i == row_loc+row_num-1:
                        ws.cell(row=i, column=j).border = thick_border

            row_num = l  # For Time Ranges of Octant's LSL
            col_num = 3
            # location of the Table
            row_loc = 1
            col_loc = 49

            for i in range(row_loc, row_loc+row_num):
                # print(ws.cell(row=i, column=50).value)
                if (ws.cell(row=i, column=50).value == " "):  # breakig if values is None( empty cell)
                    break
                for j in range(col_loc, col_num+col_loc):
                    ws.cell(row=i, column=j).border = thin_border
                    if i == row_loc+row_num-1:
                        ws.cell(row=i, column=j).border = thick_border

            sat = sat+1  # iterating to next file in input
            wb.save(out_nam)  # saving the file

    except FileNotFoundError:
        # if Input file is not found / typo in name of the file
        print("Error : Input File Not Found")



def proj_octant_gui():
    st.title("Project-2:Web Page GUI for Excel conversion")

    
    with st.sidebar:
        choice = option_menu(
            "Menu",
            ["Home", "Single File Conversion", "Bulk Conversion"],
            icons=['house', 'cloud-upload', "cloud-upload", 'gear'],
            menu_icon="cast",
            default_index=0,
            styles={
                "container": {"padding": "0!important", "background-color": "#fafafa"},
                "icon": {"color": "orange", "font-size": "20px"},
                "nav-link": {"font-size": "20px", "text-align": "left", "margin": "0px", "--hover-color": "#eee"},
                "nav-link-selected": {"background-color": "blue"},
            }
        )

    if choice == "Home":
        st.subheader("Home")

        st.markdown(""" 
        # Welcome!
        ## Select any one of the option from the menu
        ### 1.Single Excel file conversion <br> 
        ### 2.Bulk Conversion 
         """, True )
        

    elif choice == "Single File Conversion":
        st.subheader("Single File Conversion")
        data_file = st.file_uploader(
            " Upload excel file for input ", type=['xlsx'])
        mod_inp = st.number_input(
            " *Enter a Mod Value :* ", min_value=1, max_value=19996, value=5000, step=50)

        if st.button("Compute"):
            if data_file is not None:
                file_details = {"Filename": data_file.name,
                                "FileType": data_file.type, "FileSize": data_file.size}

                st.write(file_details)
                tut_07(data_file, mod_inp)
                
                st.write("Output file is saved in directory")

    elif choice == "Bulk Conversion":
        st.subheader("Bulk Conversion")
        path_inp = st.text_input("Enter a path of the folder:")
        mod_inp = st.number_input("**Enter a Mod Value :**", min_value=1, max_value=19996, value=5000, step=50)
        option=["","Default","Change the Path"]
        select_element=st.selectbox("Pick the Path for the folder to be saved",options=option)

        if(select_element=="Change the Path"):
            path_out=st.text_input("**Enter a Path of Output files to be Saved:**")
        elif(select_element=="Default"):
            path_out=r"G:\CS384\2001EE19_2022\proj2\output"   

        if st.button("Compute"):
            if path_inp is not None:
                st.markdown("### Please Wait! Converting the Input Files",True)
                tut_007(path_inp, mod_inp,path_out)
                st.markdown("## Done",True)



proj_octant_gui()


# This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
